import os
import csv
import sys

from openpyxl import load_workbook
from filter import Filter

class Reader(Filter):
    def process(self, data: None) -> dict:
        # Check if we are running as a frozen executable (e.g., packaged with PyInstaller)
        if getattr(sys, 'frozen', False):
            # If the application is running as a bundled executable, use the executable's directory
            directory = os.path.dirname(sys.executable)
        else:
            # Otherwise, we're running as a normal script, use the script's directory
            directory = os.path.dirname(os.path.realpath(__file__))

        # Get a list of all files in the directory
        all_files = os.listdir(directory)

        # Create a dictionary to store the file contents
        file_contents = {}

        # Convert each file to a list of rows and store it in the dictionary
        for file_name in all_files:
            if file_name.endswith('.csv'):
                with open(os.path.join(directory, file_name), 'r') as file:
                    reader = csv.reader(file)
                    file_contents[file_name] = list(reader)
            elif file_name == "SpendTracker.xlsx":
                with open(os.path.join(directory, file_name), 'rb') as file:
                    workbook = load_workbook(file)
                    sheet = workbook.active
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        rows.append(list(row))
                    file_contents[file_name] = rows

        return file_contents
